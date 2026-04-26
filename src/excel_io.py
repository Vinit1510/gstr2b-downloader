"""Excel I/O — read clients list, write end-of-run report, generate sample."""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config

# Required and accepted column headers (case-insensitive, trimmed)
_HEADER_ALIASES = {
    "sr no": "sr_no",
    "sr.no": "sr_no",
    "sr": "sr_no",
    "s.no": "sr_no",
    "name of client": "name",
    "client name": "name",
    "name": "name",
    "user id": "user_id",
    "userid": "user_id",
    "username": "user_id",
    "password": "password",
    "gstin": "gstin",
}


@dataclass
class Client:
    sr_no: int
    name: str
    user_id: str
    password: str
    gstin: str
    row_index: int = 0  # 1-based excel row, for reporting

    def safe_folder_name(self) -> str:
        # Keep alnum + _ only
        keep = "".join(c if c.isalnum() else "_" for c in self.name).strip("_")
        return f"{keep}_{self.gstin}" if self.gstin else keep


@dataclass
class ClientResult:
    client: Client
    status: str = "Pending"
    file_path: str = ""
    error_reason: str = ""
    started_at: str = ""
    finished_at: str = ""
    extras: dict = field(default_factory=dict)


def read_clients(path: Path) -> list[Client]:
    """Read clients from an Excel file. First row is the header."""
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        raise ValueError("Excel file is empty.")

    header_row = rows[0]
    columns: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        raw = (cell.value or "").strip().lower() if isinstance(cell.value, str) else ""
        if raw in _HEADER_ALIASES:
            columns[_HEADER_ALIASES[raw]] = idx

    missing = [k for k in ("name", "user_id", "password", "gstin") if k not in columns]
    if missing:
        raise ValueError(
            "Excel header missing required columns: "
            + ", ".join(missing)
            + ". Required: Name of Client, User ID, Password, GSTIN"
        )

    clients: list[Client] = []
    for r_idx, row in enumerate(rows[1:], start=2):
        def _get(field_key: str) -> str:
            col = columns.get(field_key)
            if col is None or col >= len(row):
                return ""
            v = row[col].value
            return str(v).strip() if v is not None else ""

        name = _get("name")
        user_id = _get("user_id")
        password = _get("password")
        gstin = _get("gstin")

        # Skip fully blank rows silently
        if not any([name, user_id, password, gstin]):
            continue

        sr_raw = _get("sr_no")
        try:
            sr_no = int(sr_raw) if sr_raw else len(clients) + 1
        except ValueError:
            sr_no = len(clients) + 1

        if not (name and user_id and password and gstin):
            # Partial row — still include but it'll fail validation in orchestrator
            pass

        clients.append(
            Client(
                sr_no=sr_no,
                name=name,
                user_id=user_id,
                password=password,
                gstin=gstin.upper(),
                row_index=r_idx,
            )
        )

    wb.close()
    return clients


def create_sample_excel(path: Path) -> None:
    """Generate a starter clients Excel with the agreed columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    headers = ["Sr No", "Name of Client", "User ID", "Password", "GSTIN"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align

    # Two dummy rows to show the format
    sample = [
        (1, "ABC Traders Pvt Ltd", "abctraders01", "DummyPass@123", "27ABCDE1234F1Z5"),
        (2, "XYZ Industries", "xyzind02", "AnotherPass!9", "29XYZAB5678K2L3"),
    ]
    for r, row in enumerate(sample, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="left" if c > 1 else "center")

    widths = [8, 36, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def write_report(path: Path, results: Iterable[ClientResult]) -> None:
    """Write end-of-run report Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-2B Report"

    headers = [
        "Sr No", "Client Name", "GSTIN", "Status",
        "File Path", "Error Reason", "Started", "Finished",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align

    status_colors = {
        "Success": "C6EFCE",
        "Already Downloaded": "DDEBF7",
        "No Data Available": "FFF2CC",
        "Failed Login": "FFC7CE",
        "Wrong Password": "FFC7CE",
        "CAPTCHA Failed": "FFEB9C",
        "Portal Error": "FFC7CE",
        "Skipped": "F2F2F2",
        "Pending": "FFFFFF",
    }

    for r, res in enumerate(results, start=2):
        c = res.client
        row_vals = [
            c.sr_no, c.name, c.gstin, res.status,
            res.file_path, res.error_reason,
            res.started_at, res.finished_at,
        ]
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(
                horizontal="left" if col in (2, 5, 6) else "center",
                vertical="center",
            )
        fill_color = status_colors.get(res.status, "FFFFFF")
        ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=fill_color)

    widths = [8, 32, 20, 22, 60, 40, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def report_filename(year: int, month: int) -> str:
    short = config.month_label(year, month)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"GSTR2B_Report_{short}_{ts}.xlsx"
